import io
import json
import logging
import openpyxl
import docx
import os
import pdfplumber
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser
from langchain_community.chat_models import ChatOllama

logger = logging.getLogger(__name__)

# --- Text and Word-Level Data Extraction ---

def extract_text_with_positions(file_content, content_type):
    """
    Extracts text and word-level positional data (bounding boxes) from a document.
    For PDFs, it uses pdfplumber. For other types, it falls back to basic text extraction
    and does not provide positional data.
    Returns a tuple: (full_text, word_map)
    - full_text: The complete extracted text of the document.
    - word_map: A list of dictionaries, where each dictionary represents a word and
                contains 'text' and its 'bbox' (bounding box) coordinates.
                Example: [{"text": "Hello", "bbox": [x0, top, x1, bottom]}, ...]
                The bbox is only available for PDFs.
    """
    logger.info(f"Extracting text and positions for content type: {content_type}")
    full_text = ""
    word_map = []

    try:
        file_stream = io.BytesIO(file_content)

        if "pdf" in content_type:
            with pdfplumber.open(file_stream) as pdf:
                for i, page in enumerate(pdf.pages):
                    page_text = page.extract_text() or ""
                    full_text += page_text + "\n"
                    
                    # Get word-level bounding boxes
                    words = page.extract_words()
                    for word in words:
                        word_map.append({
                            "text": word["text"],
                            "bbox": [word["x0"], word["top"], word["x1"], word["bottom"]],
                            "page": i + 1
                        })
        
        elif "vnd.openxmlformats-officedocument.wordprocessingml.document" in content_type: # .docx
            doc = docx.Document(file_stream)
            for para in doc.paragraphs:
                full_text += para.text + "\n"
            # Word-level data is not available for .docx with this library
        
        elif "vnd.openxmlformats-officedocument.spreadsheetml.sheet" in content_type: # .xlsx
            workbook = openpyxl.load_workbook(file_stream)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            full_text += str(cell.value) + " "
                    full_text += "\n"
            # Word-level data is not available for .xlsx

        elif "text" in content_type:
            full_text = file_content.decode('utf-8', errors='ignore')
            # Word-level data is not available for plain text
        
        else:
            logger.warning(f"Unsupported content type for positional extraction: {content_type}. Falling back to plain text.")
            full_text = file_content.decode('utf-8', errors='ignore')

    except Exception as e:
        logger.error(f"Error extracting text/positions for {content_type}: {e}", exc_info=True)
        return "", []

    if not full_text.strip():
        logger.warning(f"Could not extract any text for content_type: {content_type}")

    return full_text.strip(), word_map


# --- AI-Powered KVP Extraction with Positional Matching ---

def get_kvps_and_category_with_positions(text: str, word_map: list, examples: list = []):
    """
    Extracts Key-Value Pairs (KVPs), determines a category, and finds the bounding box
    for each KVP value using the word map.
    """
    logger.info("Initializing local LLM call to extract KVPs and category.")
    
    if not text or not text.strip():
        return {}, None, None

    system_prompt = """You are an expert document analysis AI. Analyze the document text to categorize it and extract key-value pairs (KVPs).
Return a single, valid JSON object with 'category' and 'kvps'. The 'kvps' must be a JSON object.

Example output:
{
  "category": "Invoice",
  "kvps": {
    "invoice_number": "INV-12345",
    "customer_name": "John Doe",
    "total_amount": "500.00"
  }
}"""

    try:
        llm = ChatOllama(
            base_url=os.environ.get("OLLAMA_BASE_URL"),
            model=os.environ.get("CHAT_MODEL_NAME", "phi3:mini"),
            temperature=0
        )
    except Exception as e:
        logger.error(f"Failed to initialize Ollama: {e}", exc_info=True)
        raise ConnectionError("Could not connect to the local AI model.") from e

    prompt = ChatPromptTemplate.from_messages([
        ("system", system_prompt),
        ("user", "{input_text}"),
    ])
    chain = prompt | llm | StrOutputParser()

    logger.info("Invoking LLM chain for analysis...")
    try:
        llm_response = chain.invoke({"input_text": text})
        logger.debug(f"Raw LLM response: {llm_response}")
        
        if "```json" in llm_response:
            llm_response = llm_response.split("```json")[1].split("```")[0]
        
        result = json.loads(llm_response)
        kvps_raw = result.get("kvps", {})
        category = result.get("category")
        explanation = result.get("explanation", "")

        if not isinstance(kvps_raw, dict):
            kvps_raw = {}

        # Find bounding boxes for each KVP value
        kvps_with_bbox = {k: {"value": v, "bbox": find_bbox_for_value(v, word_map)} for k, v in kvps_raw.items()}

        logger.info(f"LLM analysis successful. Category='{category}'.")
        return kvps_with_bbox, category, explanation

    except Exception as e:
        logger.error(f"Error during LLM processing or bbox mapping: {e}", exc_info=True)
        raise


def find_bbox_for_value(value: str, word_map: list):
    """
    Finds a bounding box that encapsulates the given value string by searching for
    the sequence of words in the word_map.
    Returns the combined bounding box of the word sequence or None.
    """
    if not value or not isinstance(value, str) or not word_map:
        return None

    search_words = value.strip().split()
    if not search_words:
        return None

    for i in range(len(word_map) - len(search_words) + 1):
        match = True
        for j in range(len(search_words)):
            if word_map[i + j]["text"] != search_words[j]:
                match = False
                break
        
        if match:
            # Found the sequence, now combine bounding boxes
            first_word = word_map[i]
            last_word = word_map[i + len(search_words) - 1]
            
            # Assuming words are on the same page
            if first_word["page"] != last_word["page"]:
                continue

            bbox = [
                min(w["bbox"][0] for w in word_map[i:i+len(search_words)]),
                min(w["bbox"][1] for w in word_map[i:i+len(search_words)]),
                max(w["bbox"][2] for w in word_map[i:i+len(search_words)]),
                max(w["bbox"][3] for w in word_map[i:i+len(search_words)])
            ]
            return {"page": first_word["page"], "bbox": bbox}

    return None # Not found
